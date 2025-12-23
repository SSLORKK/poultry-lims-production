from fastapi import APIRouter, Depends, Query, Response
from sqlalchemy.orm import Session, joinedload
from sqlalchemy import func, and_, case
from typing import List, Optional, Dict, Any, Set
from datetime import datetime, date
from calendar import monthrange
from dateutil.relativedelta import relativedelta
from collections import defaultdict
import json
import io
import base64

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from reportlab.lib.pagesizes import A4, letter
from reportlab.lib import colors
from reportlab.lib.units import inch
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak, Image
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.enums import TA_CENTER, TA_LEFT
from reportlab.graphics.shapes import Drawing, Rect, String, Line, Circle
from reportlab.graphics.charts.piecharts import Pie
from reportlab.graphics.charts.barcharts import VerticalBarChart
from reportlab.graphics.charts.linecharts import HorizontalLineChart
from reportlab.graphics.charts.legends import Legend
from reportlab.graphics import renderPDF
import matplotlib
matplotlib.use('Agg')  # Use non-interactive backend
import matplotlib.pyplot as plt
import matplotlib.patches as mpatches

from app.db.session import get_db
from app.models.sample import Sample
from app.models.unit import Unit
from app.models.department import Department
from app.models.pcr_data import PCRData
from app.models.serology_data import SerologyData
from app.models.microbiology_data import MicrobiologyData
from pydantic import BaseModel

router = APIRouter(prefix="/reports", tags=["reports"])


class PCRPositiveSample(BaseModel):
    farm: str
    age: Optional[str]
    house: Optional[str]
    diseases: Dict[str, str]  # disease_name -> CT value


class MicrobiologySampleType(BaseModel):
    sample_type: str
    total_count: int
    above_limit_count: int
    positive_locations: List[str]
    percentage: float


class SerologyDiseaseCount(BaseModel):
    disease_name: str
    kit_type: str
    test_count: int


class CompanyStats(BaseModel):
    company_name: str
    sample_count: int
    sub_sample_count: int
    test_count: int
    departments: Dict[str, int]
    pcr_positive_samples: Optional[List[PCRPositiveSample]] = None
    microbiology_sample_types: Optional[List[MicrobiologySampleType]] = None
    serology_diseases: Optional[List[SerologyDiseaseCount]] = None
    pcr_extraction_count: Optional[int] = None
    pcr_detection_count: Optional[int] = None


class DiseaseKitStats(BaseModel):
    disease_name: str
    kit_type: str
    test_count: int
    positive_count: int
    negative_count: int


class ReportsResponse(BaseModel):
    total_samples: int
    total_sub_samples: int
    total_tests: int
    total_positive: int
    total_negative: int
    companies: List[CompanyStats]
    diseases: List[DiseaseKitStats]
    date_range: Dict[str, str]
    department_filter: Optional[str]


class MonthDeptStats(BaseModel):
    department: str
    samples: int
    tests: int


class MonthComparisonResponse(BaseModel):
    current_month: str
    previous_month: str
    current_month_stats: List[MonthDeptStats]
    previous_month_stats: List[MonthDeptStats]
    current_month_total: Dict[str, int]
    previous_month_total: Dict[str, int]


@router.get("", response_model=ReportsResponse)
def get_comprehensive_reports(
    from_date: str = Query(..., description="Start date (YYYY-MM-DD)"),
    to_date: str = Query(..., description="End date (YYYY-MM-DD)"),
    department: Optional[str] = Query(None, description="Filter by department: PCR, SER, or MIC"),
    db: Session = Depends(get_db)
):
    """Get comprehensive reports with all analytics"""
    
    # Parse dates
    start_date = datetime.strptime(from_date, "%Y-%m-%d").date()
    end_date = datetime.strptime(to_date, "%Y-%m-%d").date()
    
    # Build sample query
    sample_query = db.query(Sample).filter(
        and_(
            Sample.date_received >= start_date,
            Sample.date_received <= end_date
        )
    )
    
    # Get all samples in date range
    all_samples = sample_query.all()
    all_sample_ids = [s.id for s in all_samples]
    
    # Build unit query with department filter
    if not all_sample_ids:
        units = []
    else:
        unit_query = db.query(Unit).join(Department).filter(
            Unit.sample_id.in_(all_sample_ids)
        ).options(
            joinedload(Unit.sample),
            joinedload(Unit.department),
            joinedload(Unit.pcr_data),
            joinedload(Unit.serology_data),
            joinedload(Unit.microbiology_data)
        )
        
        if department:
            unit_query = unit_query.filter(Department.code == department)
        
        units = unit_query.all()
    
    # Get filtered sample IDs from units (only samples that have units in the selected department)
    filtered_sample_ids = list(set(unit.sample_id for unit in units))
    
    # Calculate total samples by department logic
    try:
        if department == 'MIC':
            # For microbiology: samples = count of MIC units (1 per MIC code)
            total_samples = len(units)
            # Calculate total sub-samples for microbiology (sum of samples_number)
            total_sub_samples = sum(unit.samples_number or 0 for unit in units)
        elif department == 'PCR':
            # For PCR: samples = sum of extraction values, sub-samples = sum of samples_number
            total_samples = 0
            total_sub_samples = 0
            for unit in units:
                # Samples = extraction count
                if unit.pcr_data and unit.pcr_data.extraction:
                    total_samples += unit.pcr_data.extraction
                # Sub-samples = samples_number (samples count field)
                total_sub_samples += unit.samples_number or 0
        elif department == 'SER':
            # For Serology: samples = sum of samples_number
            total_samples = sum(unit.samples_number or 0 for unit in units)
            total_sub_samples = 0
        else:
            # For All Departments:
            # MIC: Count units (1 per unit)
            # PCR: Sum extraction values
            # SER: Sum samples_number
            total_samples = 0
            total_sub_samples = 0
            for unit in units:
                if unit.department.code == 'MIC':
                    total_samples += 1
                    total_sub_samples += unit.samples_number or 0
                elif unit.department.code == 'PCR':
                    # PCR samples = extraction count
                    if unit.pcr_data and unit.pcr_data.extraction:
                        total_samples += unit.pcr_data.extraction
                    # PCR sub-samples = samples_number
                    total_sub_samples += unit.samples_number or 0
                else:
                    total_samples += (unit.samples_number or 0)
    except Exception:
        total_samples = len(units)  # Fallback to counting units
        total_sub_samples = 0
    
    # Initialize aggregations
    company_data: Dict[str, Dict[str, Any]] = defaultdict(lambda: {
        'sample_ids': set(),
        'total_samples': 0,  # Track sum of samples_number
        'total_sub_samples': 0,  # Track sub-samples for microbiology
        'test_count': 0,
        'departments': defaultdict(int),
        'pcr_samples': [],  # For PCR pivot table
        'pcr_extraction_total': 0,
        'pcr_detection_total': 0,
        'mic_test_count': 0,  # Track microbiology tests explicitly
        'micro_sample_types': defaultdict(lambda: {
            'total': 0,
            'above_limit': 0,
            'locations': []
        }),
        'serology_diseases': defaultdict(lambda: {
            'kit_type': '',
            'count': 0
        })
    })
    
    disease_kit_data = defaultdict(lambda: {
        'test_count': 0,
        'positive_count': 0,
        'negative_count': 0
    })
    
    total_positive = 0
    total_negative = 0
    
    # Track unique diseases per sample for test counting
    sample_diseases = defaultdict(set)
    
    # Process units
    for unit in units:
        company = unit.sample.company
        dept_code = unit.department.code
        sample_id = unit.sample.id
        
        company_data[company]['sample_ids'].add(sample_id)
        try:
            if dept_code == 'MIC':
                # For microbiology: samples = count of MIC units (increment by 1 per unit)
                company_data[company]['total_samples'] += 1
                # Track sub-samples for microbiology
                company_data[company]['total_sub_samples'] += unit.samples_number or 0
            elif dept_code == 'PCR':
                # For PCR: samples = extraction count, sub-samples = samples_number
                if unit.pcr_data and unit.pcr_data.extraction:
                    company_data[company]['total_samples'] += unit.pcr_data.extraction
                # Track sub-samples for PCR (samples_number field)
                company_data[company]['total_sub_samples'] += unit.samples_number or 0
            else:
                # For other departments: samples = sum of samples_number
                company_data[company]['total_samples'] += unit.samples_number or 0
        except Exception:
            pass  # Skip if calculation causes issues
        company_data[company]['departments'][dept_code] += 1
        
        # Process PCR data
        if unit.pcr_data:
            # Add extraction and detection counts (independent of diseases_list)
            if unit.pcr_data.extraction:
                company_data[company]['pcr_extraction_total'] += unit.pcr_data.extraction
            if unit.pcr_data.detection:
                company_data[company]['pcr_detection_total'] += unit.pcr_data.detection
            
            # Process diseases if available
            if unit.pcr_data.diseases_list:
                diseases_list = unit.pcr_data.diseases_list
                if isinstance(diseases_list, str):
                    diseases_list = json.loads(diseases_list)
                
                pcr_sample_diseases = {}
                
                for disease_obj in diseases_list:
                    if isinstance(disease_obj, dict):
                        disease = disease_obj.get('disease', '')
                        kit_type = disease_obj.get('kit_type', '')
                        result = disease_obj.get('result', '')
                        ct_value = disease_obj.get('ct_value', '')
                        
                        # Skip quality control samples
                        if disease.upper() in ['POS. CONTROL', 'POS CONTROL', 'POSITIVE CONTROL', 'NEG. CONTROL', 'NEG CONTROL', 'NEGATIVE CONTROL']:
                            continue
                        
                        if disease:
                            # For PCR, don't track diseases for test counting (use detections instead)
                            # Only track diseases for non-PCR departments
                            if dept_code != 'PCR':
                                sample_diseases[sample_id].add(disease)
                            
                            # Group by disease + kit type
                            disease_kit_key = f"{disease}|||{kit_type}"
                            
                            # Use specific test count if available (new logic), otherwise default to 1
                            count_val = disease_obj.get('test_count', 1)
                            try:
                                count_val = int(count_val)
                            except (ValueError, TypeError):
                                count_val = 1
                                
                            disease_kit_data[disease_kit_key]['test_count'] += count_val
                            
                            # Count positive/negative
                            if result:
                                result_upper = result.upper().strip()
                                if result_upper in ['POS', 'POS.', 'POSITIVE', '+']:
                                    disease_kit_data[disease_kit_key]['positive_count'] += 1
                                    total_positive += 1
                                    
                                    # Store positive sample for PCR pivot table
                                    if ct_value:
                                        pcr_sample_diseases[disease] = f"CT: {ct_value}"
                                    else:
                                        pcr_sample_diseases[disease] = "POS"
                                        
                                elif result_upper in ['NEG', 'NEG.', 'NEGATIVE', '-']:
                                    disease_kit_data[disease_kit_key]['negative_count'] += 1
                                    total_negative += 1
                
                # Add to company PCR samples if any positive
                if pcr_sample_diseases:
                    company_data[company]['pcr_samples'].append({
                        'farm': unit.sample.farm or '',
                        'age': unit.sample.age or '',
                        'house': unit.house_number or '',
                        'diseases': pcr_sample_diseases
                    })
        
        # Process Serology data
        if unit.serology_data:
            diseases_list = unit.serology_data.diseases_list
            kit_type = unit.serology_data.kit_type
            tests_count = unit.serology_data.tests_count or 0
            if not kit_type:
                kit_type = 'Unknown'
            
            if isinstance(diseases_list, str):
                diseases_list = json.loads(diseases_list)
            
            # For each disease, increment count based on tests_count
            for disease_item in diseases_list:
                disease = disease_item.get('disease', 'Unknown')
                
                # Use tests_count instead of counting diseases
                disease_kit_data[disease]['test_count'] += tests_count
                
                # Track per company
                company_data[company]['serology_diseases'][disease]['kit_type'] = kit_type
                company_data[company]['serology_diseases'][disease]['count'] += tests_count
        
        # Process Microbiology data
        if unit.microbiology_data:
            diseases_list = unit.microbiology_data.diseases_list
            sample_types = unit.sample_type  # sample_type is stored in Unit, not MicrobiologyData
            
            if isinstance(diseases_list, str):
                diseases_list = json.loads(diseases_list)
            
            # Handle sample_type as array or fallback to 'Unknown'
            if sample_types is not None:
                if isinstance(sample_types, str):
                    sample_types = json.loads(sample_types)
                if isinstance(sample_types, list) and sample_types:
                    # Use first sample type or join them
                    sample_type = ', '.join(sample_types) if len(sample_types) > 1 else sample_types[0]
                else:
                    sample_type = 'Unknown'
            else:
                sample_type = 'Unknown'
            
            # Track sample type counts
            company_data[company]['micro_sample_types'][sample_type]['total'] += 1
            
            # Process diseases - microbiology diseases_list is just an array of strings
            if diseases_list:
                # Calculate tests for this unit = diseases_count * samples_number
                unit_test_count = len(diseases_list) * (unit.samples_number or 0)
                company_data[company]['mic_test_count'] += unit_test_count
                
                for disease in diseases_list:
                    if disease and isinstance(disease, str):
                        # Don't add to generic sample_diseases to avoid double counting
                        # sample_diseases[sample_id].add(disease)
                        
                        # Group by disease (no kit type for micro)
                        disease_kit_key = f"{disease}|||"
                        # For microbiology: each disease = samples_number tests
                        samples_number = unit.samples_number or 0
                        disease_kit_data[disease_kit_key]['test_count'] += samples_number
                        # Note: Microbiology doesn't store results in the diseases_list
                        # Results would need to come from a separate microbiology_coa table
    
    # Calculate total tests - sum all tests when no department filter, otherwise use department-specific logic
    try:
        if department == 'PCR':
            total_tests = sum(data['pcr_detection_total'] for data in company_data.values())
        elif department == 'SER':
            total_tests = sum(
                sum(disease_data['count'] for disease_data in data['serology_diseases'].values())
                for data in company_data.values()
            )
        elif department == 'MIC':
            # For microbiology: tests = sum of (diseases_count * samples_number) for each unit
            total_tests = 0
            for unit in units:
                if unit.microbiology_data and unit.microbiology_data.diseases_list:
                    diseases_count = len(unit.microbiology_data.diseases_list)
                    samples_number = unit.samples_number or 0
                    total_tests += diseases_count * samples_number
        elif department is None:  # All departments - sum PCR detections + Serology tests + microbiology tests + other disease tests
            pcr_tests = sum(data['pcr_detection_total'] for data in company_data.values())
            serology_tests = sum(
                sum(disease_data['count'] for disease_data in data['serology_diseases'].values())
                for data in company_data.values()
            )
            # Calculate microbiology tests
            mic_tests = 0
            for unit in units:
                if unit.microbiology_data and unit.microbiology_data.diseases_list:
                    diseases_count = len(unit.microbiology_data.diseases_list)
                    samples_number = unit.samples_number or 0
                    mic_tests += diseases_count * samples_number
            other_tests = sum(len(diseases) for diseases in sample_diseases.values())
            total_tests = pcr_tests + serology_tests + mic_tests + other_tests
        else:
            total_tests = sum(len(diseases) for diseases in sample_diseases.values())
    except Exception:
        total_tests = sum(len(diseases) for diseases in sample_diseases.values())  # Fallback
    
    # Format company stats
    companies = []
    for company, data in company_data.items():
        # For PCR, use detection total as test count
        if department == 'PCR':
            company_test_count = data['pcr_detection_total']
        elif department == 'SER':
            company_test_count = sum(
                disease_data['count'] for disease_data in data['serology_diseases'].values()
            )
        elif department == 'MIC':
            # For microbiology: calculate tests from units for this company
            company_test_count = 0
            for unit in units:
                if unit.sample.company == company and unit.microbiology_data and unit.microbiology_data.diseases_list:
                    diseases_count = len(unit.microbiology_data.diseases_list)
                    samples_number = unit.samples_number or 0
                    company_test_count += diseases_count * samples_number
        elif department is None:  # All departments
            company_test_count = data['pcr_detection_total'] + sum(
                disease_data['count'] for disease_data in data['serology_diseases'].values()
            ) + data['mic_test_count'] + sum(
                len(sample_diseases.get(sid, set()))
                for sid in data['sample_ids']
            )
        else:
            company_test_count = sum(
                len(sample_diseases.get(sid, set()))
                for sid in data['sample_ids']
            )
        
        # Prepare PCR positive samples
        pcr_positive_samples = None
        if data['pcr_samples']:
            pcr_positive_samples = [
                PCRPositiveSample(
                    farm=s['farm'],
                    age=s['age'],
                    house=s['house'],
                    diseases=s['diseases']
                )
                for s in data['pcr_samples']
            ]
        
        # Prepare Microbiology sample types
        microbiology_sample_types = None
        if data['micro_sample_types']:
            microbiology_sample_types = []
            for sample_type, type_data in data['micro_sample_types'].items():
                percentage = (type_data['above_limit'] / type_data['total'] * 100) if type_data['total'] > 0 else 0
                microbiology_sample_types.append(
                    MicrobiologySampleType(
                        sample_type=sample_type,
                        total_count=type_data['total'],
                        above_limit_count=type_data['above_limit'],
                        positive_locations=type_data['locations'],
                        percentage=round(percentage, 1)
                    )
                )
            microbiology_sample_types.sort(key=lambda x: x.total_count, reverse=True)
        
        # Prepare Serology diseases
        serology_diseases = None
        if data['serology_diseases']:
            serology_diseases = [
                SerologyDiseaseCount(
                    disease_name=disease,
                    kit_type=disease_data['kit_type'],
                    test_count=disease_data['count']
                )
                for disease, disease_data in data['serology_diseases'].items()
            ]
            serology_diseases.sort(key=lambda x: x.test_count, reverse=True)
        
        companies.append(
            CompanyStats(
                company_name=company,
                sample_count=data['total_samples'] if data['total_samples'] > 0 else len(data['sample_ids']),
                sub_sample_count=data['total_sub_samples'],
                test_count=company_test_count,
                departments=dict(data['departments']),
                pcr_positive_samples=pcr_positive_samples,
                microbiology_sample_types=microbiology_sample_types,
                serology_diseases=serology_diseases,
                pcr_extraction_count=data['pcr_extraction_total'] if data['pcr_extraction_total'] > 0 else None,
                pcr_detection_count=data['pcr_detection_total'] if data['pcr_detection_total'] > 0 else None
            )
        )
    
    # Sort by sample count (for consistent pie chart ordering)
    companies.sort(key=lambda x: x.sample_count, reverse=True)
    
    # Format disease-kit stats
    diseases = []
    for disease_kit_key, data in disease_kit_data.items():
        parts = disease_kit_key.split('|||')
        disease_name = parts[0]
        kit_type = parts[1] if len(parts) > 1 else ''
        
        diseases.append(
            DiseaseKitStats(
                disease_name=disease_name,
                kit_type=kit_type or 'Unknown',
                test_count=data['test_count'],
                positive_count=data['positive_count'],
                negative_count=data['negative_count']
            )
        )
    
    diseases.sort(key=lambda x: x.test_count, reverse=True)
    
    return ReportsResponse(
        total_samples=total_samples,
        total_sub_samples=total_sub_samples,
        total_tests=total_tests,
        total_positive=total_positive,
        total_negative=total_negative,
        companies=companies,
        diseases=diseases,
        date_range={
            "from": from_date,
            "to": to_date
        },
        department_filter=department
    )


@router.get("/export/excel")
def export_reports_excel(
    from_date: str = Query(..., description="Start date (YYYY-MM-DD)"),
    to_date: str = Query(..., description="End date (YYYY-MM-DD)"),
    department: Optional[str] = Query(None, description="Filter by department"),
    db: Session = Depends(get_db)
):
    """Export reports data to Excel"""
    
    # Get reports data
    reports_data = get_comprehensive_reports(from_date=from_date, to_date=to_date, department=department, db=db)
    
    # Create workbook
    wb = Workbook()
    if wb.active:
        wb.remove(wb.active)
    
    # Define styles
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=12)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Summary Sheet
    ws_summary = wb.create_sheet("Summary")
    ws_summary.append(["Poultry LIMS - Comprehensive Report"])
    dept_info = f" ({reports_data.department_filter})" if reports_data.department_filter else ""
    ws_summary.append(["Date Range:", f"{from_date} to {to_date}{dept_info}"])
    ws_summary.append([])
    ws_summary.append(["Metric", "Value"])
    ws_summary.append(["Total Samples", reports_data.total_samples])
    
    # Show sub-samples if relevant (mainly for Microbiology)
    if reports_data.total_sub_samples > 0:
        ws_summary.append(["Total Sub-Samples", reports_data.total_sub_samples])
        
    ws_summary.append(["Total Tests (Diseases)", reports_data.total_tests])
    
    # Only show pos/neg if not Serology
    if reports_data.department_filter != 'SER':
        ws_summary.append(["Total Positive Results", reports_data.total_positive])
        ws_summary.append(["Total Negative Results", reports_data.total_negative])
    
    ws_summary['A1'].font = Font(bold=True, size=14)
    for row in ws_summary['A4:B4']:
        for cell in row:
            cell.fill = header_fill
            cell.font = header_font
            cell.border = border
    
    # Companies Sheet (with Sub Samples)
    ws_companies = wb.create_sheet("Companies")
    ws_companies.append(["Company Type", "Total Samples", "Sub Samples", "Total Tests"])
    
    for company in reports_data.companies:
        ws_companies.append([
            company.company_name,
            company.sample_count,
            company.sub_sample_count,
            company.test_count
        ])
    
    for row in ws_companies['A1:D1']:
        for cell in row:
            cell.fill = header_fill
            cell.font = header_font
            cell.border = border
    
    # Samples Distribution Sheet
    ws_samples_dist = wb.create_sheet("Samples Distribution")
    ws_samples_dist.append(["Company", "Sample Count", "Percentage"])
    total_samples = sum(c.sample_count for c in reports_data.companies)
    for company in reports_data.companies:
        percentage = (company.sample_count / total_samples * 100) if total_samples > 0 else 0
        ws_samples_dist.append([
            company.company_name,
            company.sample_count,
            f"{percentage:.1f}%"
        ])
    
    for row in ws_samples_dist['A1:C1']:
        for cell in row:
            cell.fill = header_fill
            cell.font = header_font
            cell.border = border
    
    # Tests Distribution Sheet
    ws_tests_dist = wb.create_sheet("Tests Distribution")
    ws_tests_dist.append(["Company", "Test Count", "Percentage"])
    total_tests_all = sum(c.test_count for c in reports_data.companies)
    for company in reports_data.companies:
        percentage = (company.test_count / total_tests_all * 100) if total_tests_all > 0 else 0
        ws_tests_dist.append([
            company.company_name,
            company.test_count,
            f"{percentage:.1f}%"
        ])
    
    for row in ws_tests_dist['A1:C1']:
        for cell in row:
            cell.fill = header_fill
            cell.font = header_font
            cell.border = border
    
    # Month Comparison Sheet
    try:
        month_comparison = get_month_comparison(db=db)
        ws_comparison = wb.create_sheet("Monthly Comparison")
        ws_comparison.append([f"Monthly Comparison: {month_comparison.current_month} vs {month_comparison.previous_month}"])
        ws_comparison.append([])
        ws_comparison.append(["Department", f"{month_comparison.current_month} Samples", f"{month_comparison.previous_month} Samples", 
                             "Samples Change", f"{month_comparison.current_month} Tests", f"{month_comparison.previous_month} Tests", "Tests Change"])
        
        for i, curr in enumerate(month_comparison.current_month_stats):
            prev = month_comparison.previous_month_stats[i] if i < len(month_comparison.previous_month_stats) else None
            sample_diff = curr.samples - (prev.samples if prev else 0)
            test_diff = curr.tests - (prev.tests if prev else 0)
            ws_comparison.append([
                curr.department,
                curr.samples,
                prev.samples if prev else 0,
                f"+{sample_diff}" if sample_diff > 0 else str(sample_diff),
                curr.tests,
                prev.tests if prev else 0,
                f"+{test_diff}" if test_diff > 0 else str(test_diff)
            ])
        
        # Add totals row
        total_sample_diff = month_comparison.current_month_total['samples'] - month_comparison.previous_month_total['samples']
        total_test_diff = month_comparison.current_month_total['tests'] - month_comparison.previous_month_total['tests']
        ws_comparison.append([
            "TOTAL",
            month_comparison.current_month_total['samples'],
            month_comparison.previous_month_total['samples'],
            f"+{total_sample_diff}" if total_sample_diff > 0 else str(total_sample_diff),
            month_comparison.current_month_total['tests'],
            month_comparison.previous_month_total['tests'],
            f"+{total_test_diff}" if total_test_diff > 0 else str(total_test_diff)
        ])
        
        # Style the comparison sheet
        comparison_header_fill = PatternFill(start_color="6366f1", end_color="6366f1", fill_type="solid")
        ws_comparison['A1'].font = Font(bold=True, size=14)
        for row in ws_comparison['A3:G3']:
            for cell in row:
                cell.fill = comparison_header_fill
                cell.font = header_font
                cell.border = border
    except Exception as e:
        print(f"Error adding month comparison to Excel: {e}")
    
    # Disease-Kit Type Sheet
    ws_diseases = wb.create_sheet("Disease-Kit Breakdown")
    headers = ["Disease Name", "Kit Type", "Test Count"]
    if reports_data.department_filter != 'SER':
        headers.extend(["Positive Count", "Negative Count", "Positive Rate %"])
    ws_diseases.append(headers)
    
    for disease in reports_data.diseases:
        row_data = [
            disease.disease_name,
            disease.kit_type,
            disease.test_count
        ]
        if reports_data.department_filter != 'SER':
            positive_rate = (disease.positive_count / disease.test_count * 100) if disease.test_count > 0 else 0
            row_data.extend([
                disease.positive_count,
                disease.negative_count,
                f"{positive_rate:.1f}%"
            ])
        ws_diseases.append(row_data)
    
    for row in ws_diseases['A1:F1']:
        for cell in row:
            cell.fill = header_fill
            cell.font = header_font
            cell.border = border
    
    # Auto-adjust column widths for all sheets
    all_sheets = [ws_summary, ws_companies, ws_samples_dist, ws_tests_dist, ws_diseases]
    # Add comparison sheet if it exists
    if 'Monthly Comparison' in wb.sheetnames:
        all_sheets.append(wb['Monthly Comparison'])
    
    for ws in all_sheets:
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
    
    # Save to bytes
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    dept_suffix = f"_{department}" if department else ""
    filename = f"LIMS_Report_{from_date}_to_{to_date}{dept_suffix}.xlsx"
    
    return Response(
        content=output.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


def create_reportlab_pie_chart(data: List[Dict], title: str, width: float = 450, height: float = 250) -> Drawing:
    """Create a professional pie chart using ReportLab's drawing library"""
    d = Drawing(width, height)
    
    # Filter out zero values and get top 6
    filtered_data = [item for item in data[:6] if item.get('value', 0) > 0]
    if not filtered_data:
        d.add(String(width/2, height/2, 'No Data', textAnchor='middle', fontSize=14))
        return d
    
    # Create pie chart
    pie = Pie()
    pie.x = 80
    pie.y = 30
    pie.width = 150
    pie.height = 150
    pie.data = [item['value'] for item in filtered_data]
    pie.labels = [f"{item['label'][:15]}" for item in filtered_data]
    
    # Professional color scheme
    chart_colors = [
        colors.HexColor('#3b82f6'),  # Blue
        colors.HexColor('#10b981'),  # Green
        colors.HexColor('#a855f7'),  # Purple
        colors.HexColor('#f59e0b'),  # Amber
        colors.HexColor('#ef4444'),  # Red
        colors.HexColor('#8b5cf6'),  # Violet
    ]
    
    for i in range(len(filtered_data)):
        pie.slices[i].fillColor = chart_colors[i % len(chart_colors)]
        pie.slices[i].strokeColor = colors.white
        pie.slices[i].strokeWidth = 2
        pie.slices[i].popout = 3
    
    pie.slices.strokeWidth = 1
    pie.slices.fontName = 'Helvetica'
    pie.slices.fontSize = 8
    pie.sideLabels = True
    pie.sideLabelsOffset = 0.1
    
    d.add(pie)
    
    # Add title
    d.add(String(width/2, height - 20, title, textAnchor='middle', fontSize=12, fontName='Helvetica-Bold'))
    
    # Add legend on the right side
    legend = Legend()
    legend.x = 280
    legend.y = height - 60
    legend.dx = 8
    legend.dy = 8
    legend.fontName = 'Helvetica'
    legend.fontSize = 9
    legend.boxAnchor = 'nw'
    legend.columnMaximum = 6
    legend.strokeWidth = 0.5
    legend.strokeColor = colors.HexColor('#e5e7eb')
    legend.deltax = 75
    legend.deltay = 10
    legend.autoXPadding = 5
    legend.yGap = 0
    legend.dxTextSpace = 5
    legend.alignment = 'right'
    legend.dividerLines = 1|2|4
    legend.dividerOffsY = 4.5
    legend.subCols.rpad = 30
    
    total = sum(item['value'] for item in filtered_data)
    legend.colorNamePairs = [
        (chart_colors[i % len(chart_colors)], f"{item['label'][:12]} ({item['value']}) - {item['value']/total*100:.1f}%")
        for i, item in enumerate(filtered_data)
    ]
    
    d.add(legend)
    
    return d


def create_reportlab_bar_chart(data: Dict, title: str, width: float = 450, height: float = 280) -> Drawing:
    """Create a professional grouped bar chart using ReportLab"""
    d = Drawing(width, height)
    
    labels = data.get('labels', [])
    if not labels:
        d.add(String(width/2, height/2, 'No Data', textAnchor='middle', fontSize=14))
        return d
    
    current_samples = data.get('current_samples', [])
    previous_samples = data.get('previous_samples', [])
    current_tests = data.get('current_tests', [])
    previous_tests = data.get('previous_tests', [])
    
    # Create bar chart
    bc = VerticalBarChart()
    bc.x = 60
    bc.y = 50
    bc.height = height - 100
    bc.width = width - 120
    bc.data = [current_samples, previous_samples, current_tests, previous_tests]
    bc.strokeColor = colors.black
    bc.valueAxis.valueMin = 0
    bc.valueAxis.labels.fontName = 'Helvetica'
    bc.valueAxis.labels.fontSize = 9
    bc.categoryAxis.labels.boxAnchor = 'ne'
    bc.categoryAxis.labels.dx = -8
    bc.categoryAxis.labels.dy = -2
    bc.categoryAxis.labels.angle = 30
    bc.categoryAxis.labels.fontName = 'Helvetica-Bold'
    bc.categoryAxis.labels.fontSize = 10
    bc.categoryAxis.categoryNames = labels
    
    # Bar colors
    bc.bars[0].fillColor = colors.HexColor('#6366f1')  # Current Samples
    bc.bars[1].fillColor = colors.HexColor('#a5b4fc')  # Previous Samples
    bc.bars[2].fillColor = colors.HexColor('#06b6d4')  # Current Tests
    bc.bars[3].fillColor = colors.HexColor('#67e8f9')  # Previous Tests
    
    bc.barWidth = 8
    bc.groupSpacing = 15
    bc.barSpacing = 2
    
    d.add(bc)
    
    # Add title
    d.add(String(width/2, height - 15, title, textAnchor='middle', fontSize=11, fontName='Helvetica-Bold'))
    
    # Add legend
    legend = Legend()
    legend.x = width - 180
    legend.y = height - 35
    legend.fontName = 'Helvetica'
    legend.fontSize = 8
    legend.boxAnchor = 'nw'
    legend.columnMaximum = 2
    legend.alignment = 'right'
    
    current_month = data.get('current_month', 'Current')
    previous_month = data.get('previous_month', 'Previous')
    legend.colorNamePairs = [
        (colors.HexColor('#6366f1'), f'{current_month} Samples'),
        (colors.HexColor('#a5b4fc'), f'{previous_month} Samples'),
        (colors.HexColor('#06b6d4'), f'{current_month} Tests'),
        (colors.HexColor('#67e8f9'), f'{previous_month} Tests'),
    ]
    
    d.add(legend)
    
    return d


@router.get("/export/pdf")
def export_reports_pdf(
    from_date: str = Query(..., description="Start date (YYYY-MM-DD)"),
    to_date: str = Query(..., description="End date (YYYY-MM-DD)"),
    department: Optional[str] = Query(None, description="Filter by department"),
    db: Session = Depends(get_db)
):
    """Export reports data to PDF organized by departments with charts"""
    
    def create_chart_image(chart_type: str, data: Dict, title: str, dept_color: str = '#3b82f6') -> io.BytesIO:
        """Create a chart and return as BytesIO image"""
        fig, ax = plt.subplots(figsize=(7, 4))
        
        if chart_type == 'bar_horizontal':
            labels = [item['label'][:30] for item in data['items'][:10]]
            values = [item['value'] for item in data['items'][:10]]
            y_pos = range(len(labels))
            ax.barh(y_pos, values, color=dept_color, alpha=0.8)
            ax.set_yticks(y_pos)
            ax.set_yticklabels(labels, fontsize=9)
            ax.set_xlabel(data.get('xlabel', 'Count'), fontsize=10)
            ax.set_title(title, fontsize=12, fontweight='bold')
            ax.invert_yaxis()
            
        elif chart_type == 'bar_vertical':
            labels = [item['label'][:15] for item in data['items'][:10]]
            values = [item['value'] for item in data['items'][:10]]
            x_pos = range(len(labels))
            
            # Use colorful palette matching frontend
            chart_colors = ['#3b82f6', '#10b981', '#a855f7', '#f59e0b', '#ef4444', '#8b5cf6', '#ec4899', '#14b8a6']
            bar_colors = [chart_colors[i % len(chart_colors)] for i in range(len(values))]
            
            ax.bar(x_pos, values, color=bar_colors, alpha=0.8)
            ax.set_xticks(x_pos)
            ax.set_xticklabels(labels, rotation=45, ha='right', fontsize=8)
            ax.set_ylabel(data.get('ylabel', 'Count'), fontsize=10)
            ax.set_title(title, fontsize=12, fontweight='bold')
            
        elif chart_type == 'stacked_bar':
            labels = [item['label'][:20] for item in data['items'][:10]]
            positive = [item.get('positive', 0) for item in data['items'][:10]]
            negative = [item.get('negative', 0) for item in data['items'][:10]]
            x_pos = range(len(labels))
            ax.bar(x_pos, positive, label='Positive', color='#ef4444', alpha=0.8)
            ax.bar(x_pos, negative, bottom=positive, label='Negative', color='#10b981', alpha=0.8)
            ax.set_xticks(x_pos)
            ax.set_xticklabels(labels, rotation=45, ha='right', fontsize=8)
            ax.set_ylabel('Test Count', fontsize=10)
            ax.set_title(title, fontsize=12, fontweight='bold')
            ax.legend()
        
        elif chart_type == 'pie':
            # Enhanced pie chart with better styling
            fig.set_size_inches(8, 5)
            ax.clear()
            
            labels = [item['label'][:18] for item in data['items'][:6]]
            values = [item['value'] for item in data['items'][:6]]
            chart_colors = ['#3b82f6', '#10b981', '#a855f7', '#f59e0b', '#ef4444', '#8b5cf6']
            
            # Filter out zero values
            filtered_data = [(l, v, c) for l, v, c in zip(labels, values, chart_colors) if v > 0]
            if filtered_data:
                labels, values, pie_colors = zip(*filtered_data)
                
                # Create pie chart with better styling
                wedges, texts, autotexts = ax.pie(
                    values, 
                    labels=None,  # We'll use legend instead
                    autopct='%1.1f%%',
                    colors=pie_colors,
                    startangle=90,
                    pctdistance=0.6,
                    explode=[0.02] * len(values),  # Slight separation
                    shadow=True,
                    wedgeprops={'edgecolor': 'white', 'linewidth': 2}
                )
                
                # Style the percentage text
                for autotext in autotexts:
                    autotext.set_fontsize(9)
                    autotext.set_fontweight('bold')
                    autotext.set_color('white')
                
                # Add legend on the right side
                ax.legend(wedges, labels, title="Companies", loc="center left", 
                         bbox_to_anchor=(1, 0, 0.5, 1), fontsize=9)
                
                ax.set_title(title, fontsize=14, fontweight='bold', pad=15, color='#1f2937')
            else:
                ax.text(0.5, 0.5, 'No Data', ha='center', va='center', fontsize=14)
                ax.set_title(title, fontsize=14, fontweight='bold')
        
        elif chart_type == 'line':
            # Line chart for month comparison trends
            labels = data.get('labels', [])
            current_samples = data.get('current_samples', [])
            previous_samples = data.get('previous_samples', [])
            current_tests = data.get('current_tests', [])
            previous_tests = data.get('previous_tests', [])
            
            x = range(len(labels))
            
            # Plot lines with markers
            ax.plot(x, current_samples, 'o-', linewidth=2.5, markersize=8, 
                   label=f'{data.get("current_month", "Current")} Samples', color='#6366f1')
            ax.plot(x, previous_samples, 's--', linewidth=2.5, markersize=8,
                   label=f'{data.get("previous_month", "Previous")} Samples', color='#a5b4fc')
            ax.plot(x, current_tests, '^-', linewidth=2.5, markersize=8,
                   label=f'{data.get("current_month", "Current")} Tests', color='#06b6d4')
            ax.plot(x, previous_tests, 'd--', linewidth=2.5, markersize=8,
                   label=f'{data.get("previous_month", "Previous")} Tests', color='#14b8a6')
            
            ax.set_xticks(x)
            ax.set_xticklabels(labels, fontsize=11, fontweight='bold')
            ax.set_ylabel('Count', fontsize=11)
            ax.set_title(title, fontsize=13, fontweight='bold', color='#1f2937')
            ax.legend(loc='upper right', fontsize=9, framealpha=0.9)
            ax.grid(True, linestyle='--', alpha=0.7)
            ax.set_facecolor('#f8fafc')
        
        elif chart_type == 'grouped_bar':
            # For month comparison chart
            labels = data.get('labels', [])
            current_samples = data.get('current_samples', [])
            previous_samples = data.get('previous_samples', [])
            current_tests = data.get('current_tests', [])
            previous_tests = data.get('previous_tests', [])
            
            x = range(len(labels))
            width = 0.2
            
            ax.bar([i - 1.5*width for i in x], current_samples, width, label=f'{data.get("current_month", "Current")} Samples', color='#6366f1')
            ax.bar([i - 0.5*width for i in x], previous_samples, width, label=f'{data.get("previous_month", "Previous")} Samples', color='#a5b4fc')
            ax.bar([i + 0.5*width for i in x], current_tests, width, label=f'{data.get("current_month", "Current")} Tests', color='#06b6d4')
            ax.bar([i + 1.5*width for i in x], previous_tests, width, label=f'{data.get("previous_month", "Previous")} Tests', color='#67e8f9')
            
            ax.set_xticks(x)
            ax.set_xticklabels(labels, fontsize=10)
            ax.set_ylabel('Count', fontsize=10)
            ax.set_title(title, fontsize=12, fontweight='bold')
            ax.legend(loc='upper right', fontsize=8)
        
        plt.tight_layout()
        img_buffer = io.BytesIO()
        plt.savefig(img_buffer, format='png', dpi=100, bbox_inches='tight')
        img_buffer.seek(0)
        plt.close(fig)
        return img_buffer
    
    # Get reports data for each department or specified department
    departments_to_process = []
    if department:
        departments_to_process = [department]
    else:
        # Get all departments from database
        all_depts = db.query(Department).all()
        departments_to_process = [d.code for d in all_depts]
    
    # Create PDF
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4, rightMargin=30, leftMargin=30, topMargin=50, bottomMargin=30)
    
    # Styles
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=20,
        textColor=colors.HexColor('#366092'),
        spaceAfter=20,
        alignment=TA_CENTER,
        fontName='Helvetica-Bold'
    )
    dept_title_style = ParagraphStyle(
        'DeptTitle',
        parent=styles['Heading1'],
        fontSize=16,
        spaceAfter=15,
        spaceBefore=10,
        alignment=TA_LEFT,
        fontName='Helvetica-Bold'
    )
    heading_style = ParagraphStyle(
        'CustomHeading',
        parent=styles['Heading2'],
        fontSize=13,
        textColor=colors.HexColor('#366092'),
        spaceAfter=10,
        spaceBefore=10,
        fontName='Helvetica-Bold'
    )
    
    # Build document
    elements = []
    
    # Main Title
    elements.append(Paragraph("Central Poultry Laboratory - Comprehensive Report", title_style))
    elements.append(Paragraph(f"Date Range: {from_date} to {to_date}", styles['Normal']))
    elements.append(Spacer(1, 20))
    
    # Get overall data for pie charts (all departments or filtered)
    overall_data = get_comprehensive_reports(from_date=from_date, to_date=to_date, department=department, db=db)
    
    # Add Samples Distribution Pie Chart using ReportLab
    if overall_data.companies:
        elements.append(Paragraph("Samples Distribution by Company", heading_style))
        samples_pie_data = [{'label': c.company_name, 'value': c.sample_count} for c in overall_data.companies if c.sample_count > 0]
        if samples_pie_data:
            pie_drawing = create_reportlab_pie_chart(samples_pie_data, 'Samples Distribution', width=480, height=220)
            elements.append(pie_drawing)
            elements.append(Spacer(1, 25))
    
    # Add Tests Distribution Pie Chart using ReportLab
    if overall_data.companies:
        elements.append(Paragraph("Tests Distribution by Company", heading_style))
        tests_pie_data = [{'label': c.company_name, 'value': c.test_count} for c in overall_data.companies if c.test_count > 0]
        if tests_pie_data:
            pie_drawing = create_reportlab_pie_chart(tests_pie_data, 'Tests Distribution', width=480, height=220)
            elements.append(pie_drawing)
            elements.append(Spacer(1, 25))
    
    # Add Month Comparison Charts using ReportLab
    try:
        month_comparison = get_month_comparison(db=db)
        if month_comparison.current_month_stats:
            elements.append(Paragraph(f"Monthly Comparison: {month_comparison.current_month} vs {month_comparison.previous_month}", heading_style))
            
            comparison_data = {
                'labels': [s.department for s in month_comparison.current_month_stats],
                'current_samples': [s.samples for s in month_comparison.current_month_stats],
                'previous_samples': [month_comparison.previous_month_stats[i].samples if i < len(month_comparison.previous_month_stats) else 0 for i in range(len(month_comparison.current_month_stats))],
                'current_tests': [s.tests for s in month_comparison.current_month_stats],
                'previous_tests': [month_comparison.previous_month_stats[i].tests if i < len(month_comparison.previous_month_stats) else 0 for i in range(len(month_comparison.current_month_stats))],
                'current_month': month_comparison.current_month,
                'previous_month': month_comparison.previous_month
            }
            
            # Add Bar Chart using ReportLab
            bar_drawing = create_reportlab_bar_chart(comparison_data, f'Monthly Comparison: {month_comparison.current_month} vs {month_comparison.previous_month}', width=500, height=280)
            elements.append(bar_drawing)
            elements.append(Spacer(1, 15))
            elements.append(Spacer(1, 15))
            
            # Add Line Chart
            elements.append(Paragraph("Trend Comparison (Line Chart)", heading_style))
            img_buffer_line = create_chart_image('line', comparison_data, f'Line Chart: {month_comparison.current_month} vs {month_comparison.previous_month}')
            img_line = Image(img_buffer_line, width=6.5*inch, height=3.5*inch)
            elements.append(img_line)
            elements.append(Spacer(1, 15))
            
            # Add comparison summary table
            comparison_table_data = [
                ['Department', f'{month_comparison.current_month} Samples', f'{month_comparison.previous_month} Samples', f'{month_comparison.current_month} Tests', f'{month_comparison.previous_month} Tests']
            ]
            for i, curr in enumerate(month_comparison.current_month_stats):
                prev = month_comparison.previous_month_stats[i] if i < len(month_comparison.previous_month_stats) else None
                comparison_table_data.append([
                    curr.department,
                    str(curr.samples),
                    str(prev.samples if prev else 0),
                    str(curr.tests),
                    str(prev.tests if prev else 0)
                ])
            comparison_table_data.append([
                'Total',
                str(month_comparison.current_month_total['samples']),
                str(month_comparison.previous_month_total['samples']),
                str(month_comparison.current_month_total['tests']),
                str(month_comparison.previous_month_total['tests'])
            ])
            
            comp_table = Table(comparison_table_data, colWidths=[1.2*inch, 1.2*inch, 1.2*inch, 1.2*inch, 1.2*inch])
            comp_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#6366f1')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, -1), 9),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 10),
                ('BACKGROUND', (0, -1), (-1, -1), colors.HexColor('#e0e7ff')),
                ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
                ('GRID', (0, 0), (-1, -1), 1, colors.black),
                ('ROWBACKGROUNDS', (0, 1), (-1, -2), [colors.white, colors.HexColor('#f5f5f5')])
            ]))
            elements.append(comp_table)
            elements.append(Spacer(1, 20))
    except Exception as e:
        print(f"Error adding month comparison: {e}")
    
    elements.append(PageBreak())
    
    # Process each department
    for dept_code in departments_to_process:
        # Convert to string to avoid SQLAlchemy Column type issues
        dept_code_str = str(dept_code)
        
        # Get department-specific data
        dept_data = get_comprehensive_reports(from_date=from_date, to_date=to_date, department=dept_code_str, db=db)
        
        if dept_data.total_samples == 0:
            continue
        
        # Department color coding
        dept_colors = {
            'PCR': ('#3b82f6', colors.HexColor('#3b82f6')),
            'SER': ('#10b981', colors.HexColor('#10b981')),
            'MIC': ('#a855f7', colors.HexColor('#a855f7'))
        }
        chart_color, table_color = dept_colors.get(dept_code_str, ('#366092', colors.HexColor('#366092')))
        
        # Department Title
        dept_name = {'PCR': 'PCR Department', 'SER': 'Serology Department', 'MIC': 'Microbiology Department'}.get(dept_code_str, dept_code_str)
        elements.append(Paragraph(dept_name, dept_title_style))
        elements.append(Spacer(1, 10))
        
        # Summary Statistics
        elements.append(Paragraph("Summary Statistics", heading_style))
        summary_data = [
            ['Metric', 'Value'],
            ['Total Samples', str(dept_data.total_samples)],
        ]
        
        # Add sub-samples if present
        if dept_data.total_sub_samples > 0:
            summary_data.append(['Total Sub-Samples', str(dept_data.total_sub_samples)])
            
        summary_data.append(['Total Tests', str(dept_data.total_tests)])
        
        if dept_code_str != 'SER':
            summary_data.extend([
                ['Positive Results', str(dept_data.total_positive)],
                ['Negative Results', str(dept_data.total_negative)],
                ['Positivity Rate', f"{(dept_data.total_positive / dept_data.total_tests * 100):.1f}%" if dept_data.total_tests > 0 else '0%']
            ])
        
        summary_table = Table(summary_data, colWidths=[3*inch, 2*inch])
        summary_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), table_color),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 11),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        elements.append(summary_table)
        elements.append(Spacer(1, 20))
        
        # Disease Chart
        if dept_data.diseases:
            elements.append(Paragraph(f"{dept_name} - Disease Distribution", heading_style))
            
            # Create chart data
            chart_data = {
                'items': [{
                    'label': d.disease_name,
                    'value': d.test_count,
                    'positive': d.positive_count,
                    'negative': d.negative_count
                } for d in dept_data.diseases[:10]]
            }
            
            if dept_code_str != 'SER':
                # Use colorful vertical bar chart for all departments
                img_buffer = create_chart_image('bar_vertical', chart_data, 'Diseases Tested', chart_color)
            else:
                # For Serology, used to use horizontal, now unify to vertical for consistency? 
                # User asked for "SAME STYLE". Frontend uses Vertical for SER too now (Consolidated).
                img_buffer = create_chart_image('bar_vertical', chart_data, 'Diseases Tested', chart_color)
            
            img = Image(img_buffer, width=5.5*inch, height=3*inch)
            elements.append(img)
            elements.append(Spacer(1, 15))
        
        # Companies Table
        if dept_data.companies:
            elements.append(Paragraph("Companies", heading_style))
            company_data = [['Company', 'Samples', 'Tests']]
            for company in dept_data.companies[:15]:
                company_data.append([
                    company.company_name[:35],
                    str(company.sample_count),
                    str(company.test_count)
                ])
            
            company_table = Table(company_data, colWidths=[3*inch, 1*inch, 1*inch])
            company_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), table_color),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (1, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, -1), 9),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 10),
                ('BACKGROUND', (0, 1), (-1, -1), colors.white),
                ('GRID', (0, 0), (-1, -1), 1, colors.black),
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.lightgrey])
            ]))
            elements.append(company_table)
            elements.append(Spacer(1, 15))
        
        # Disease-Kit Type Detailed Table
        elements.append(Paragraph("Disease-Kit Type Details", heading_style))
        if dept_code_str != 'SER':
            disease_headers = ['Disease', 'Kit', 'Tests', 'Pos', 'Neg', 'Rate']
            col_widths = [2*inch, 1.3*inch, 0.6*inch, 0.6*inch, 0.6*inch, 0.7*inch]
        else:
            disease_headers = ['Disease', 'Kit Type', 'Tests']
            col_widths = [2.5*inch, 2*inch, 1*inch]
        
        disease_data = [disease_headers]
        for disease in dept_data.diseases[:25]:
            row = [
                disease.disease_name[:25],
                disease.kit_type[:18],
                str(disease.test_count)
            ]
            if dept_code_str != 'SER':
                pos_rate = (disease.positive_count / disease.test_count * 100) if disease.test_count > 0 else 0
                row.extend([
                    str(disease.positive_count),
                    str(disease.negative_count),
                    f"{pos_rate:.1f}%"
                ])
            disease_data.append(row)
        
        disease_table = Table(disease_data, colWidths=col_widths)
        disease_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), table_color),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (1, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 8),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 10),
            ('BACKGROUND', (0, 1), (-1, -1), colors.white),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.lightgrey])
        ]))
        elements.append(disease_table)
        
        # Add page break between departments (except for last one)
        if dept_code_str != departments_to_process[-1]:
            elements.append(PageBreak())
    
    # Build PDF
    doc.build(elements)
    buffer.seek(0)
    
    dept_suffix = f"_{department}" if department else "_All_Departments"
    filename = f"LIMS_Report_{from_date}_to_{to_date}{dept_suffix}.pdf"
    
    return Response(
        content=buffer.getvalue(),
        media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


@router.get("/month-comparison", response_model=MonthComparisonResponse)
def get_month_comparison(
    db: Session = Depends(get_db)
):
    """Get comparison of samples and tests between current and previous month by department"""
    today = date.today()
    
    # Current month range
    current_month_start = today.replace(day=1)
    current_month_end = today.replace(day=monthrange(today.year, today.month)[1])
    
    # Previous month range
    prev_month_date = today - relativedelta(months=1)
    previous_month_start = prev_month_date.replace(day=1)
    previous_month_end = prev_month_date.replace(day=monthrange(prev_month_date.year, prev_month_date.month)[1])
    
    # Get all departments
    departments = db.query(Department).all()
    
    def get_stats_for_period(start_date: date, end_date: date) -> tuple:
        """Get samples and tests count by department for a date range"""
        dept_stats = []
        total_samples = 0
        total_tests = 0
        
        for dept in departments:
            # Count samples
            sample_count = db.query(func.count(func.distinct(Sample.id))).join(
                Unit, Sample.id == Unit.sample_id
            ).filter(
                Unit.department_id == dept.id,
                Sample.date_received >= start_date,
                Sample.date_received <= end_date
            ).scalar() or 0
            
            # Count tests based on department type
            test_count = 0
            if dept.code == 'PCR':
                # For PCR, count detection tests from PCRData
                test_count = db.query(func.coalesce(func.sum(PCRData.detection), 0)).join(
                    Unit, PCRData.unit_id == Unit.id
                ).join(
                    Sample, Unit.sample_id == Sample.id
                ).filter(
                    Unit.department_id == dept.id,
                    Sample.date_received >= start_date,
                    Sample.date_received <= end_date
                ).scalar() or 0
            elif dept.code == 'SER':
                test_count = db.query(func.count(SerologyData.id)).join(
                    Unit, SerologyData.unit_id == Unit.id
                ).join(
                    Sample, Unit.sample_id == Sample.id
                ).filter(
                    Unit.department_id == dept.id,
                    Sample.date_received >= start_date,
                    Sample.date_received <= end_date
                ).scalar() or 0
            elif dept.code == 'MIC':
                test_count = db.query(func.count(MicrobiologyData.id)).join(
                    Unit, MicrobiologyData.unit_id == Unit.id
                ).join(
                    Sample, Unit.sample_id == Sample.id
                ).filter(
                    Unit.department_id == dept.id,
                    Sample.date_received >= start_date,
                    Sample.date_received <= end_date
                ).scalar() or 0
            
            dept_stats.append(MonthDeptStats(
                department=dept.code,
                samples=sample_count,
                tests=int(test_count)
            ))
            total_samples += sample_count
            total_tests += int(test_count)
        
        return dept_stats, {"samples": total_samples, "tests": total_tests}
    
    current_stats, current_total = get_stats_for_period(current_month_start, current_month_end)
    previous_stats, previous_total = get_stats_for_period(previous_month_start, previous_month_end)
    
    return MonthComparisonResponse(
        current_month=current_month_start.strftime("%B %Y"),
        previous_month=previous_month_start.strftime("%B %Y"),
        current_month_stats=current_stats,
        previous_month_stats=previous_stats,
        current_month_total=current_total,
        previous_month_total=previous_total
    )
